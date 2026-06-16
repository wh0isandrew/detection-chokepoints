#!/usr/bin/env python3
"""
build_domain_monthly.py
Recompute the CLEAN per-domain ClickFix behavioural trend from Carson Williams'
ClickFix Hunter export and write it into _data/clickgrab_trends.yml.

WHY this source (see docs/DECISIONS.md #012): the page's delivery-chain analysis
(cradle families, evasion, inline payloads) must come from Carson's export, which
records the ACTUAL clipboard command per domain (clean), NOT from MHaggis ClickGrab's
per-site web crawls (whose site-level cradle/evasion classification is dominated by
'Likely Safe' lure pages and UI-string base64 — a judge panel measured ~93-99% noise).
Carson's export = the real ClickFix command per domain, so classifying it is honest.

Carson export schema (sheet 'Data'): domain | url | timestamp | commandline
Output (into _data/clickgrab_trends.yml, other sections preserved):
  domain_monthly[]        per-month technique prevalence (drives the charts)
  domain_cradles_total    cumulative cradle counts (drives the stat cards)
  domain_evasion_totals   cumulative evasion counts (base64 / hex_xor / inline)
  meta.total_domains      unique ClickFix domains in the export

Usage:
  python scripts/build_domain_monthly.py [path/to/clickfix-domains-all-YYYY-MM-DD.xlsx]
Default input: newest clickfix-domains-all-*.xlsx in ~/Downloads.
"""

import glob
import os
import re
import sys
from collections import OrderedDict, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import yaml

REPO_ROOT  = Path(__file__).parent.parent
TRENDS_YML = REPO_ROOT / "_data" / "clickgrab_trends.yml"
DEFAULT_GLOB = str(Path(os.path.expanduser("~")) / "Downloads" / "clickfix-domains-all-*.xlsx")

# ---------------------------------------------------------------------------
# Classification — applied to the real clipboard command string per domain.
# Counts are PREVALENCE (non-exclusive): a domain is counted under every technique
# its command uses, matching the page's "msiexec 87% (669/767)" framing.
# ---------------------------------------------------------------------------

CRADLE_RE = OrderedDict([
    ("iwr",       re.compile(r"\biwr\b|invoke-webrequest", re.I)),
    ("irm",       re.compile(r"\birm\b|invoke-restmethod", re.I)),
    ("webclient", re.compile(r"webclient|downloadstring|downloadfile", re.I)),
    ("curl",      re.compile(r"\bcurl\b", re.I)),
    ("msiexec",   re.compile(r"\bmsiexec\b", re.I)),
    ("mshta",     re.compile(r"\bmshta\b", re.I)),
    ("vbs",       re.compile(r"winhttp|\bwscript\b|createobject\(|\.vbs\b", re.I)),
])
EVASION_RE = OrderedDict([
    ("hex_xor",   re.compile(r"-bxor", re.I)),
    # base64: FromBase64String, or -EncodedCommand and its valid PowerShell prefix
    # abbreviations (-e .. -encodedcommand). The base64-blob lookahead disambiguates
    # -e from -ExecutionPolicy, which is followed by a word, not a blob (DECISIONS #013).
    ("base64",    re.compile(
        r"frombase64string"
        r"|-e(?:n(?:c(?:o(?:d(?:e(?:d(?:c(?:o(?:m(?:m(?:a(?:n(?:d)?)?)?)?)?)?)?)?)?)?)?)?)?"
        r"(?=\s+[\"']?[A-Za-z0-9+/]{16,})", re.I)),
])
# no_url means "no REMOTE FETCH", not merely "no https?:// URL": also catch
# single-slash http:/, ftp, UNC/WebDAV (\\host\.., \\IP@port\DavWWWRoot), and
# scheme-less bare-IPv4 fetches (irm/iwr/curl <ip>) (DECISIONS #013).
REMOTE_FETCH_RE = re.compile(
    r"https?:/+|ftp://|\\\\[\w.$@-]|\b(?:\d{1,3}\.){3}\d{1,3}\b", re.I)

CRADLE_KEYS  = list(CRADLE_RE.keys())
EVASION_KEYS = list(EVASION_RE.keys()) + ["no_url"]


def classify_commandline(cmd: str) -> dict:
    """Return {technique: bool} for one clipboard command. `no_url` = an inline
    payload (the command fetches nothing remote)."""
    flags = {}
    for k, rx in CRADLE_RE.items():
        flags[k] = bool(rx.search(cmd))
    for k, rx in EVASION_RE.items():
        flags[k] = bool(rx.search(cmd))
    flags["no_url"] = not bool(REMOTE_FETCH_RE.search(cmd))
    return flags


# ---------------------------------------------------------------------------
# Read the export
# ---------------------------------------------------------------------------

def _pick_export(arg: str | None) -> Path:
    if arg:
        return Path(arg)
    matches = sorted(glob.glob(DEFAULT_GLOB))
    if not matches:
        print(f"ERROR: no clickfix-domains-all-*.xlsx found in {DEFAULT_GLOB}", file=sys.stderr)
        sys.exit(1)
    # Filenames carry the export date — newest by name == latest snapshot.
    return Path(matches[-1])


def read_rows(path: Path) -> list:
    """Return [(domain, timestamp, commandline)] from the export's Data sheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip().lower() if c is not None else "" for c in next(it)]
    idx = {name: header.index(name) for name in ("domain", "timestamp", "commandline") if name in header}
    if not {"domain", "timestamp", "commandline"} <= idx.keys():
        print(f"ERROR: export missing expected columns; got {header}", file=sys.stderr)
        sys.exit(1)
    out = []
    for r in it:
        dom = r[idx["domain"]]
        ts  = r[idx["timestamp"]]
        cmd = r[idx["commandline"]]
        if dom:
            out.append((str(dom).strip().lower(), ts, str(cmd or "")))
    wb.close()
    return out


def _month(ts) -> str:
    if isinstance(ts, datetime):
        return ts.strftime("%Y-%m")
    s = str(ts or "")
    return s[:7] if len(s) >= 7 and s[4] == "-" else ""


# ---------------------------------------------------------------------------
# Aggregate
# ---------------------------------------------------------------------------

def build(rows: list) -> dict:
    # Dedup by domain (an export row == a unique ClickFix domain); keep the first
    # observation's month + command so re-counts are stable.
    seen = {}
    for dom, ts, cmd in rows:
        if dom not in seen:
            seen[dom] = (_month(ts), cmd)

    months = defaultdict(lambda: {"n": 0, **{k: 0 for k in CRADLE_KEYS}, **{k: 0 for k in EVASION_KEYS}})
    cradles_total = {k: 0 for k in CRADLE_KEYS}
    evasion_total = {k: 0 for k in EVASION_KEYS}

    for dom, (mon, cmd) in seen.items():
        if not mon:
            continue
        flags = classify_commandline(cmd)
        m = months[mon]
        m["n"] += 1
        for k in CRADLE_KEYS:
            m[k] += flags[k]; cradles_total[k] += flags[k]
        for k in EVASION_KEYS:
            m[k] += flags[k]; evasion_total[k] += flags[k]

    domain_monthly = []
    for mon in sorted(months):
        m = months[mon]
        n = m["n"]
        entry = {"month": mon, "n": n}
        for k in CRADLE_KEYS:
            entry[k] = m[k]
        entry["hex_xor"] = m["hex_xor"]
        entry["base64"]  = m["base64"]
        entry["no_url"]  = m["no_url"]
        entry["no_url_pct"] = round(100 * m["no_url"] / n, 1) if n else 0.0
        domain_monthly.append(entry)

    return {
        "domain_monthly": domain_monthly,
        "domain_cradles_total": cradles_total,
        "domain_evasion_totals": evasion_total,
        "total_domains": len(seen),
    }


# ---------------------------------------------------------------------------
# Write
# ---------------------------------------------------------------------------

def main() -> None:
    export = _pick_export(sys.argv[1] if len(sys.argv) > 1 else None)
    print(f"Reading Carson export: {export.name}")
    rows = read_rows(export)
    result = build(rows)
    dm = result["domain_monthly"]
    print(f"  {result['total_domains']} unique domains over {len(dm)} months "
          f"({dm[0]['month']} .. {dm[-1]['month']})")

    if not TRENDS_YML.exists():
        print(f"ERROR: {TRENDS_YML} missing", file=sys.stderr)
        sys.exit(1)
    data = yaml.safe_load(TRENDS_YML.read_text(encoding="utf-8"))

    data["domain_monthly"] = dm
    data["domain_cradles_total"] = result["domain_cradles_total"]
    data["domain_evasion_totals"] = result["domain_evasion_totals"]
    meta = data.setdefault("meta", {})
    meta["total_domains"] = result["total_domains"]
    meta["domain_source"] = export.name
    meta["domain_updated"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    header = (
        "# Generated by scripts/analyze_clickgrab.py (volume) + build_domain_monthly.py (behaviour).\n"
        "# domain_monthly / domain_*_total are the CLEAN per-domain command classification from\n"
        "# Carson's ClickFix Hunter export and drive the page charts/cards (see DECISIONS #012).\n\n"
    )
    body = yaml.safe_dump(data, sort_keys=False, allow_unicode=True, default_flow_style=False, width=4096)
    TRENDS_YML.write_text(header + body, encoding="utf-8")

    print(f"  cradles: {result['domain_cradles_total']}")
    print(f"  evasion: {result['domain_evasion_totals']}")
    print(f"Written: {TRENDS_YML}")


if __name__ == "__main__":
    main()
